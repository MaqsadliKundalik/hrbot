from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from database.models import AdminsResume, TeacherResume, TgUser, QuizAnswers, Quizs
import io
import os
from collections import defaultdict
from PIL import Image as PILImage


HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
ALT_FILL = PatternFill(start_color="F2F7FF", end_color="F2F7FF", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Rasm o'lchami (piksel)
IMG_SIZE = 55
ROW_HEIGHT = 45   # piksel (Excel units ≈ pt, 1pt ≈ 0.75px → 45px ≈ 34pt)
ROW_HEIGHT_PT = 45
IMG_COL_WIDTH = 8  # Excel column width units


def _style_header_row(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _style_data_row(ws, row: int, col_count: int):
    fill = ALT_FILL if row % 2 == 0 else None
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        if fill:
            cell.fill = fill
        cell.alignment = LEFT
        cell.border = THIN_BORDER


def _set_column_widths(ws, widths: list):
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 28


def _get_photo_path(tg_id: int) -> str | None:
    """statics/photos/{tg_id}.jpg yo'lini qaytaradi, mavjud bo'lsa."""
    path = f"statics/photos/{tg_id}.jpg"
    return path if os.path.exists(path) else None


def _make_thumb(path: str, size: int = IMG_SIZE) -> io.BytesIO | None:
    """Rasmni kvadrat thumbnail qilib BytesIO qaytaradi."""
    try:
        img = PILImage.open(path).convert("RGB")
        img.thumbnail((size, size), PILImage.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=85)
        buf.seek(0)
        return buf
    except Exception:
        return None


def _insert_photo(ws, row: int, col: int, tg_id: int):
    """Rasm mavjud bo'lsa Excelga joylashtiradi."""
    path = _get_photo_path(tg_id)
    if not path:
        return
    thumb = _make_thumb(path)
    if not thumb:
        return
    img = XLImage(thumb)
    img.width = IMG_SIZE
    img.height = IMG_SIZE
    col_letter = get_column_letter(col)
    # Rasmni katakka to'g'ri joylashtirish uchun offset (2pt)
    img.anchor = f"{col_letter}{row}"
    ws.add_image(img)


def _safe_sheet_title(title: str) -> str:
    for ch in r'\/*?:[]\r\n':
        title = title.replace(ch, " ")
    return title[:31]


async def generate_report() -> io.BytesIO:
    wb = Workbook()
    default_sheet = wb.active

    admin_headers = [
        "№", "Rasm", "Ism Familiya", "Telegram ID", "Telefon",
        "Filial", "Tug'ilgan sana", "Tug'ilgan joyi", "Yashash manzili", "Xorijiy til", "Til darajasi",
        "Tajriba", "Ish vaqti",
        "Oxirgi ish joyi", "Ketish sababi", "Tel (ish joyi)",
        "Nega bizni tanladi", "Bizni qayerdan topdi", "Ariza sanasi",
    ]
    admin_widths = [4, IMG_COL_WIDTH, 24, 16, 18, 15, 14, 20, 20, 16, 14, 14, 14, 22, 22, 18, 25, 20, 18]

    admins = await AdminsResume.all().prefetch_related("user").order_by("job", "id")
    admin_groups: dict[str, list] = defaultdict(list)
    for resume in admins:
        admin_groups[resume.job].append(resume)

    for job, resumes in admin_groups.items():
        ws = wb.create_sheet(title=job)
        ws.append(admin_headers)
        _style_header_row(ws, 1, len(admin_headers))
        _set_column_widths(ws, admin_widths)

        for idx, resume in enumerate(resumes, start=1):
            user: TgUser = resume.user
            phones = ", ".join(v for v in user.phone_numbers.values() if v) if user.phone_numbers else "—"
            excel_row = idx + 1

            row_data = [
                idx,
                "",   # Rasm ustuni
                user.full_name,
                user.tg_id,
                phones,
                user.branch or "—",
                str(user.birth_date) if user.birth_date else "—",
                user.born_address or "—",
                user.live_address or "—",
                resume.foreign_language,
                resume.foreign_language_level,
                resume.experience,
                resume.working_time,
                resume.last_work_place,
                resume.why_leave_work,
                resume.last_work_place_phone,
                resume.why_choice_us or "—",
                user.where_find_us or "—",
                resume.created_at.strftime("%d.%m.%Y %H:%M") if resume.created_at else "—",
            ]
            ws.append(row_data)
            _style_data_row(ws, excel_row, len(admin_headers))
            ws.row_dimensions[excel_row].height = ROW_HEIGHT_PT
            _insert_photo(ws, excel_row, 2, user.tg_id)

    teacher_headers = [
        "№", "Rasm", "Ism Familiya", "Telegram ID", "Telefon",
        "Filial", "Tug'ilgan sana", "Tug'ilgan joyi", "Yashash manzili", "Tajriba", "Lavozim", "Ish vaqti", "Maosh (so'm)",
        "Sertifikatlar",
        "Oxirgi ish joyi", "Ketish sababi", "Tel (ish joyi)",
        "Nega bizni tanladi", "Bizni qayerdan topdi", "Ariza sanasi",
    ]
    teacher_widths = [4, IMG_COL_WIDTH, 24, 16, 18, 15, 14, 20, 20, 13, 13, 13, 16, 32, 22, 22, 18, 25, 20, 18]

    teachers = await TeacherResume.all().prefetch_related("user").order_by("subject", "id")
    teacher_groups: dict[str, list] = defaultdict(list)
    for resume in teachers:
        teacher_groups[resume.subject].append(resume)

    for subject, resumes in teacher_groups.items():
        ws = wb.create_sheet(title=subject)
        ws.append(teacher_headers)
        _style_header_row(ws, 1, len(teacher_headers))
        _set_column_widths(ws, teacher_widths)

        for idx, resume in enumerate(resumes, start=1):
            user: TgUser = resume.user
            phones = ", ".join(v for v in user.phone_numbers.values() if v) if user.phone_numbers else "—"
            certs = (
                "\n".join(
                    f"{s.get('name', '?')} — {s.get('ball', '?')}"
                    for s in resume.sertificates
                )
                if resume.sertificates
                else "—"
            )
            excel_row = idx + 1

            row_data = [
                idx,
                "",   # Rasm
                user.full_name,
                user.tg_id,
                phones,
                user.branch or "—",
                str(user.birth_date) if user.birth_date else "—",
                user.born_address or "—",
                user.live_address or "—",
                resume.experience,
                resume.position,
                resume.working_time,
                resume.salary,
                certs,
                resume.last_work_place,
                resume.why_leave_work,
                resume.last_work_place_phone,
                resume.why_choice_us or "—",
                user.where_find_us or "—",
                resume.created_at.strftime("%d.%m.%Y %H:%M") if resume.created_at else "—",
            ]
            ws.append(row_data)
            _style_data_row(ws, excel_row, len(teacher_headers))
            ws.row_dimensions[excel_row].height = ROW_HEIGHT_PT
            _insert_photo(ws, excel_row, 2, user.tg_id)

    ws_quizzes = wb.create_sheet(title="Test natijalari")
    quiz_headers = ["№", "Telegram ID", "Ism Familiya", "Fan", "To'g'ri javoblar"]
    quiz_widths = [5, 18, 25, 25, 18]
    ws_quizzes.append(quiz_headers)
    _style_header_row(ws_quizzes, 1, len(quiz_headers))
    _set_column_widths(ws_quizzes, quiz_widths)

    quiz_answers = (
        await QuizAnswers.all()
        .prefetch_related("user", "quiz", "quiz__subject")
        .order_by("id")
    )
    for idx, answer in enumerate(quiz_answers, start=1):
        user: TgUser = answer.user
        quiz: Quizs = answer.quiz
        subject_name = quiz.subject.name if quiz.subject else "—"
        ws_quizzes.append([idx, user.tg_id, user.full_name, subject_name, answer.correct_answers])
        _style_data_row(ws_quizzes, idx + 1, len(quiz_headers))

    wb.remove(default_sheet)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
