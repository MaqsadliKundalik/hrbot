import os
from openpyxl import load_workbook

async def get_test_questions(file_path: str) -> list[dict]:
    """
    Excel faylidan test savollarini o'qiydi va validatsiya qiladi.
    
    Fayl strukturasi:
    A: Savol
    B-E: Variantlar (4 ta)
    F: To'g'ri javob
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Fayl topilmadi: {file_path}")
    
    if not file_path.endswith('.xlsx'):
        raise ValueError("Fayl formati noto'g'ri. Faqat .xlsx fayllari qabul qilinadi.")

    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        raise ValueError(f"Excel faylini ochishda xatolik yuz berdi: {str(e)}")
        
    sheet = wb.active
    questions = []
    
    # row index for error reporting (starts from 2 because min_row=2)
    for idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=6, values_only=True), start=2):
        # Agar savol ustuni bo'sh bo'lsa, bu qatorni tashlab ketamiz
        if not row[0]:
            continue
            
        question_text = str(row[0]).strip()
        options = [str(opt).strip() if opt is not None else "" for opt in row[1:5]]
        
        if any(not opt for opt in options):
            raise ValueError(f"Qatorda ({idx}) variantlar to'liq emas. 4 ta variant bo'lishi shart.")

        try:
            correct_num = int(row[5]) if row[5] is not None else None
        except (ValueError, TypeError):
            raise ValueError(f"Qatorda ({idx}) to'g'ri variant raqami noto'g'ri formatda. Faqat 1 dan 4 gacha raqam bo'lishi kerak.")

        if correct_num is None or not (1 <= correct_num <= 4):
            raise ValueError(f"Qatorda ({idx}) to'g'ri variant raqami 1 va 4 oralig'ida bo'lishi shart.")

        answer = options[correct_num - 1]

        if not answer:
            raise ValueError(f"Qatorda ({idx}) tanlangan to'g'ri variant (variant {correct_num}) bo'sh bo'lishi mumkin emas.")

        questions.append({
            "question": question_text,
            "options": options,
            "answer": answer
        })
        
    if not questions:
        raise ValueError("Faylda birorta ham yaroqli savol topilmadi.")
        
    return questions